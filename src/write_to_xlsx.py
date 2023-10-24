from write_to_xlsx_argv import *              # This file should contains variable: workbook_name, L, T
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Series, Reference
import os

def column_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def table_label( row, col ): 
    s = column_string( col ) + str(row) 
    return s

def write_result( Routing_Policy, MST_Policy, L, T ):

    # set style 
    thin            = Side(border_style="thin", color="000000")
    centor_align    = Alignment( horizontal='center', vertical='center' )

    # write headers
    ws.cell( row = 1, column = 2, value = 'Wire Length (unit)' ).alignment=centor_align
    ws.cell( row = 1, column = 7, value = 'Time (microsecond)' ).alignment=centor_align
    for i in range( len(MST_Policy) ):
        ws.cell( row = i + 3, column = 1, value = MST_Policy[i] ).alignment=centor_align
    for rep in range( 2 ):
        for i in range( len(Routing_Policy) ):
            ws.cell( row = 2, column = 2 + i + rep * len(Routing_Policy), value = Routing_Policy[i] ).alignment=centor_align
    # write wire length data    
    for k in range( len(L) ):
        i = k // len(Routing_Policy)
        j = k %  len(Routing_Policy)
        ws.cell( row = 3 + i, column = 2 + j, value = L[k] ).alignment=centor_align
    # write time data    
    for k in range( len(T) ):
        i = k // len(Routing_Policy)
        j = k %  len(Routing_Policy)
        ws.cell( row = 3 + i, column = 7 + j, value = T[k] ).alignment=centor_align

    # merge cells
    ws.merge_cells( start_row = 1, start_column = 2, end_row = 1, end_column = 6  )
    ws.merge_cells( start_row = 1, start_column = 7, end_row = 1, end_column = 11 )

    # set border style
    for i in range( len(MST_Policy) + 2 ):
        ws.cell( row = i + 1, column = 1 ).border = Border( right = thin )
        ws.cell( row = i + 1, column = 1 + len(Routing_Policy) ).border = Border( right = thin )
        ws.cell( row = i + 1, column = 1 + 2*len(Routing_Policy) ).border = Border( right = thin ) 
    for i in range( 2*len(Routing_Policy) + 1 ):
        if ( i + 1 == 1 or i + 1 == 1 + len(Routing_Policy) or i + 1 == 1 + 2*len(Routing_Policy) ):
            ws.cell( row = 3, column = i + 1 ).border = Border( right = thin, top = thin ) 
        else:
            ws.cell( row = 2, column = i + 1 ).border = Border( bottom = thin ) 

def write_normalization( Routing_Policy, MST_Policy, L, T, offset ):

    # set style 
    thin            = Side(border_style="thin", color="000000")
    centor_align    = Alignment( horizontal='center', vertical='center' )

    # write headers
    ws.cell( row = 1 + offset, column = 2, value = 'Wire Length (normalized)' ).alignment=centor_align
    for i in range( len(MST_Policy) ):
        ws.cell( row = i + 3 + offset, column = 1, value = MST_Policy[i] ).alignment=centor_align
    for i in range( len(Routing_Policy) ):
        ws.cell( row = 2 + offset, column = 2 + i, value = Routing_Policy[i] ).alignment=centor_align
    # write wire length data    
    for k in range( len(L) ):
        if ( k == 0 ):
            ws[table_label(3+offset, 2)] = 1
            ws[table_label(3+offset, 2)].alignment=centor_align
            continue  
        i = k // len(Routing_Policy)
        j = k %  len(Routing_Policy)
        ws[table_label(3 + i + offset,2 + j)] = "="+table_label(3+i,2+j)+"/"+table_label(3,2) 
        ws[table_label(3 + i + offset,2 + j)].alignment=centor_align

    # merge cells
    ws.merge_cells( start_row = 1 + offset, start_column = 2, end_row = 1 + offset, end_column = 6  )

    # set border style
    for i in range( len(MST_Policy) + 2 ):
        ws.cell( row = i + 1 + offset, column = 1 ).border = Border( right = thin )
        ws.cell( row = i + 1 + offset, column = 1 + len(Routing_Policy) ).border = Border( right = thin )
    for i in range( len(Routing_Policy) + 1 ):
        if ( i + 1 == 1 or i + 1 == 1 + len(Routing_Policy) or i + 1 == 1 + 2*len(Routing_Policy) ):
            ws.cell( row = 3 + offset, column = i + 1 ).border = Border( right = thin, top = thin ) 
        else:
            ws.cell( row = 2 + offset, column = i + 1 ).border = Border( bottom = thin ) 


def add_image( Routing_Policy, MST_Policy, P ):
    # set style 
    centor_align    = Alignment( horizontal='center', vertical='center' )

    for k in range( len(P) ):
        img = Image( P[k] + '.png' )
        img.width = 1200 // 5
        img.height = 1000 // 5  
        i = k // len(Routing_Policy)
        j = k %  len(Routing_Policy)
        ws.add_image(img, table_label( 20 + 12*i, 1 + 3*j ))
        ws.cell( row = 19 + 12*i, column = 1 + 3*j, value = MST_Policy[i]+"+"+Routing_Policy[j] ).alignment=centor_align
        ws.merge_cells( start_row = 19 + 12*i, start_column = 1 + 3*j, end_row = 19 + 12*i, end_column =  1 + 3*j + 2 )

if __name__ == '__main__':

    wb = Workbook()
    ws = wb.active

    write_result( Routing_Policy, MST_Policy, L, T )
    write_normalization( Routing_Policy, MST_Policy, L, T, 3+len(MST_Policy) )
    add_image( Routing_Policy, MST_Policy, P )

    # change column width 
    for i in range( 2*len(Routing_Policy) + 1 ):
        ws.column_dimensions[column_string(i+1)].width = 10 

    # Save the output file
    wb.save( workbook_name )

    # remove all PNGs
    for k in range( len(P) ):
        if ( P[k] != './src/NoResult' ):
            os.remove( P[k] + '.png' )

